"""
Extract CDO 2026 Budget data from Excel into a structured JSON file
for the web dashboard to consume.
"""

import json
import os
from typing import Any

import openpyxl

EXCEL_PATH = (
    "/Users/TLorino/Library/CloudStorage/OneDrive-DigitalRealty/"
    "Documents/CDO Ops/Budget/CDO Budget - 2026-02-08.xlsx"
)
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "src", "data", "budget.json")

MONTHS = [
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
]

QUARTERS = ["Q1", "Q2", "Q3", "Q4"]


def _safe_float(val: Any) -> float:
    """Convert a cell value to float, defaulting to 0."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def extract() -> dict:
    """Parse the Excel workbook and return structured budget data."""
    wb_values = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    wb_formulas = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
    ws = wb_values["2026 Budget"]
    ws_f = wb_formulas["2026 Budget"]

    # ---- Metadata ----
    metadata = {
        "title": "CDO 2026 Budget",
        "file": "CDO Budget - 2026-02-08.xlsx",
        "sheet": "2026 Budget",
        "table_name": "CDOBudget",
        "total_columns": ws.max_column,
        "total_line_items": 0,
        "extracted_at": str(__import__("datetime").date.today()),
    }

    # ---- Column mapping (1-indexed) ----
    # A=1 Category, B=2 Event Type, C=3 Source Fund, D=4 Expense Type,
    # E=5 Vendor, F=6 Description, G=7 ITBO Line, H=8 SOW, I=9 ITBO, J=10 PO
    # K=11 Approved FY26, L=12 Committed FY26
    # M-X (13-24) Budget monthly Jan-Dec
    # Y-AB (25-28) Budget quarterly Q1-Q4
    # AC=29 Budget FY26
    # AD-AO (30-41) Forecast monthly Jan-Dec
    # AP-AS (42-45) Forecast quarterly Q1-Q4
    # AT=46 Forecast FY26
    # AU-BF (47-58) Actual monthly Jan-Dec
    # BG-BJ (59-62) Actual quarterly Q1-Q4
    # BK=63 Actual FY26
    # BL-BW (64-75) BvF monthly Jan-Dec
    # BX-CA (76-79) BvF quarterly Q1-Q4
    # CB=80 BvF FY26
    # CC-CN (81-92) BvA monthly Jan-Dec
    # CO-CR (93-96) BvA quarterly Q1-Q4
    # CS=97 BvA FY26

    line_items = []
    for row_idx in range(3, 52):  # rows 3-51
        category = ws.cell(row_idx, 1).value
        event_type = ws.cell(row_idx, 2).value
        if not event_type:
            continue

        item = {
            "row": row_idx,
            "category": str(category or ""),
            "event_type": str(event_type or ""),
            "source_fund": str(ws.cell(row_idx, 3).value or ""),
            "expense_type": str(ws.cell(row_idx, 4).value or ""),
            "vendor": str(ws.cell(row_idx, 5).value or ""),
            "description": str(ws.cell(row_idx, 6).value or ""),
            "itbo_line": str(ws.cell(row_idx, 7).value or ""),
            "sow": str(ws.cell(row_idx, 8).value or ""),
            "itbo": str(ws.cell(row_idx, 9).value or ""),
            "po": str(ws.cell(row_idx, 10).value or ""),
        }

        item["budget_fy26"] = _safe_float(ws.cell(row_idx, 11).value)
        item["committed_fy26"] = _safe_float(ws.cell(row_idx, 12).value)

        # Budget monthly: 13-24 (M-X)
        item["budget_monthly"] = [_safe_float(ws.cell(row_idx, 13 + m).value) for m in range(12)]
        # Budget quarterly: 25-28 (Y-AB)
        item["budget_quarterly"] = [_safe_float(ws.cell(row_idx, 25 + q).value) for q in range(4)]
        # Budget FY: 29 (AC)
        item["budget_fy"] = _safe_float(ws.cell(row_idx, 29).value)

        # Forecast monthly: 30-41 (AD-AO)
        item["forecast_monthly"] = [_safe_float(ws.cell(row_idx, 30 + m).value) for m in range(12)]
        # Forecast quarterly: 42-45 (AP-AS)
        item["forecast_quarterly"] = [_safe_float(ws.cell(row_idx, 42 + q).value) for q in range(4)]
        # Forecast FY: 46 (AT)
        item["forecast_fy"] = _safe_float(ws.cell(row_idx, 46).value)

        # Actual monthly: 47-58 (AU-BF)
        item["actual_monthly"] = [_safe_float(ws.cell(row_idx, 47 + m).value) for m in range(12)]
        # Actual quarterly: 59-62 (BG-BJ)
        item["actual_quarterly"] = [_safe_float(ws.cell(row_idx, 59 + q).value) for q in range(4)]
        # Actual FY: 63 (BK)
        item["actual_fy"] = _safe_float(ws.cell(row_idx, 63).value)

        # BvF monthly: 64-75 (BL-BW)
        item["bvf_monthly"] = [_safe_float(ws.cell(row_idx, 64 + m).value) for m in range(12)]
        # BvF FY: 80 (CB)
        item["bvf_fy"] = _safe_float(ws.cell(row_idx, 80).value)

        # BvA monthly: 81-92 (CC-CN)
        item["bva_monthly"] = [_safe_float(ws.cell(row_idx, 81 + m).value) for m in range(12)]
        # BvA FY: 97 (CS)
        item["bva_fy"] = _safe_float(ws.cell(row_idx, 97).value)

        # Committed formula for audit display
        l_formula = ws_f.cell(row_idx, 12).value
        item["committed_formula"] = str(l_formula) if l_formula else ""

        line_items.append(item)

    metadata["total_line_items"] = len(line_items)

    # ---- Totals row (row 52) ----
    totals = {
        "budget_fy26": _safe_float(ws.cell(52, 11).value),
        "committed_fy26": _safe_float(ws.cell(52, 12).value),
        "budget_monthly": [_safe_float(ws.cell(52, 13 + m).value) for m in range(12)],
        "budget_quarterly": [_safe_float(ws.cell(52, 25 + q).value) for q in range(4)],
        "budget_fy": _safe_float(ws.cell(52, 29).value),
        "forecast_monthly": [_safe_float(ws.cell(52, 30 + m).value) for m in range(12)],
        "forecast_quarterly": [_safe_float(ws.cell(52, 42 + q).value) for q in range(4)],
        "forecast_fy": _safe_float(ws.cell(52, 46).value),
        "actual_monthly": [_safe_float(ws.cell(52, 47 + m).value) for m in range(12)],
        "actual_quarterly": [_safe_float(ws.cell(52, 59 + q).value) for q in range(4)],
        "actual_fy": _safe_float(ws.cell(52, 63).value),
        "bvf_fy": _safe_float(ws.cell(52, 80).value),
        "bva_fy": _safe_float(ws.cell(52, 97).value),
    }

    # ---- Aggregations ----
    by_source_fund: dict[str, float] = {}
    by_expense_type: dict[str, float] = {}
    by_category: dict[str, float] = {}
    actuals_detail: list[dict] = []

    for item in line_items:
        fund = item["source_fund"] or "Unspecified"
        by_source_fund[fund] = by_source_fund.get(fund, 0) + item["committed_fy26"]

        etype = item["expense_type"] or "Unspecified"
        by_expense_type[etype] = by_expense_type.get(etype, 0) + item["committed_fy26"]

        cat = item["category"] or "Other"
        by_category[cat] = by_category.get(cat, 0) + item["committed_fy26"]

        if item["actual_fy"] > 0:
            actuals_detail.append({
                "category": item["category"],
                "vendor": item["vendor"],
                "description": item["description"],
                "actual_fy": item["actual_fy"],
            })

    # ---- Formula audit findings ----
    audit_findings = [
        {
            "id": 1,
            "severity": "high",
            "title": "Over-Committed Budget Categories",
            "detail": (
                "Contingent Labor D1: Approved $1.589M but SOWs total $1.646M "
                "(Committed remainder = -$56,562). EDA Budget Capex: Approved "
                "$313K but SOWs total $873K (Committed remainder = -$559,722). "
                "More is committed than the approved allocation."
            ),
        },
        {
            "id": 2,
            "severity": "medium",
            "title": "Flat Forecast Distribution",
            "detail": (
                "Forecast monthly values use Committed / 12 (or / 11 for late "
                "starts). Categories like Travel, Training, and Marketing have "
                "lumpy real-world spending patterns. Consider seasonalizing these "
                "forecasts to improve monthly accuracy."
            ),
        },
        {
            "id": 3,
            "severity": "medium",
            "title": "Forecast Not Yet Rolling Forward",
            "detail": (
                "As actuals are booked, Forecast should update to reflect "
                "remaining spend: (Committed - YTD Actuals) / remaining months. "
                "Currently Forecast still uses the original Committed / 12 "
                "even for months with actuals."
            ),
        },
        {
            "id": 4,
            "severity": "low",
            "title": "Floating-Point Rounding Artifacts",
            "detail": (
                "Some totals show sub-penny rounding differences due to "
                "division-based formulas. Wrapping key formulas in ROUND() "
                "would eliminate this."
            ),
        },
        {
            "id": 5,
            "severity": "info",
            "title": "Budget = Approved Plan (Fixed Baseline)",
            "detail": (
                "Budget monthly columns correctly reference Approved FY26 / 12 "
                "on Budget rows and are blank on SOW rows. This establishes the "
                "approved plan as an immutable baseline for variance analysis."
            ),
        },
        {
            "id": 6,
            "severity": "info",
            "title": "Forecast = Committed Best Estimate",
            "detail": (
                "Forecast monthly columns reference Committed FY26 / 12 on both "
                "Budget remainder rows and SOW rows. Budget vs Forecast variance "
                "now shows the gap between the approved plan and current committed "
                "spend expectations."
            ),
        },
        {
            "id": 7,
            "severity": "info",
            "title": "Unapproved Items Separated",
            "detail": (
                "EDA Projects rows use 'Unapproved Budget' and 'Unapproved SOW' "
                "event types. These are excluded from approved budget rollups and "
                "shown separately in the dashboard as pending approval."
            ),
        },
    ]

    # ---- Recommendations ----
    recommendations = [
        {
            "id": 1,
            "priority": "high",
            "title": "Fix Over-Committed Categories",
            "detail": (
                "Contingent Labor D1 and EDA Budget Capex have negative Committed "
                "remainders. Either increase the Approved allocation or reduce SOW "
                "commitments to bring them back in balance."
            ),
            "impact": "Prevents unplanned budget overruns",
        },
        {
            "id": 2,
            "priority": "high",
            "title": "Roll Forecast Forward with Actuals",
            "detail": (
                "As monthly actuals are booked, update Forecast for remaining "
                "months to (Committed - YTD Actuals) / remaining months. This "
                "makes Forecast a living best estimate rather than a static plan."
            ),
            "impact": "Accurate real-time financial visibility",
        },
        {
            "id": 3,
            "priority": "medium",
            "title": "Seasonalize Forecast Allocations",
            "detail": (
                "Replace flat Committed/12 with realistic monthly profiles for "
                "lumpy categories: Travel (conference months), Training (Q1/Q3), "
                "project ramp-ups (phased milestones)."
            ),
            "impact": "More accurate monthly Forecast vs Actual comparisons",
        },
        {
            "id": 4,
            "priority": "medium",
            "title": "Add Percentage Variance Columns",
            "detail": (
                "Add (Budget - Actual) / Budget percentage columns alongside "
                "the existing dollar variance. Percentages contextualize the "
                "magnitude of variances across different-sized line items."
            ),
            "impact": "Better executive-level variance analysis",
        },
        {
            "id": 5,
            "priority": "medium",
            "title": "Add Conditional Formatting in Excel",
            "detail": (
                "Apply green/yellow/red color coding in the spreadsheet: red for "
                "negative Committed remainders, traffic-light colors for Budget "
                "vs Actual thresholds (>10% under = green, 0-10% = yellow, "
                "over = red)."
            ),
            "impact": "At-a-glance health indicators for budget status",
        },
        {
            "id": 6,
            "priority": "low",
            "title": "Add Forecast vs Actual Variance Section",
            "detail": (
                "Add a third variance section (Forecast - Actual) to measure "
                "forecasting accuracy. This completes the analysis triangle: "
                "Budget vs Actual, Budget vs Forecast, Forecast vs Actual."
            ),
            "impact": "Measures and improves forecasting accuracy over time",
        },
        {
            "id": 7,
            "priority": "low",
            "title": "Add Year-over-Year Comparison",
            "detail": (
                "Pull 2025 Budget actuals into a YoY delta view. The 2025 Budget "
                "tab already exists with historical data."
            ),
            "impact": "Trend analysis and planning improvement",
        },
    ]

    return {
        "metadata": metadata,
        "months": MONTHS,
        "quarters": QUARTERS,
        "line_items": line_items,
        "totals": totals,
        "by_source_fund": by_source_fund,
        "by_expense_type": by_expense_type,
        "by_category": by_category,
        "actuals_detail": actuals_detail,
        "audit_findings": audit_findings,
        "recommendations": recommendations,
    }


def main() -> None:
    data = extract()
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w") as f:
        json.dump(data, f, indent=2, default=str)
    approved = data["totals"]["budget_fy26"]
    committed = data["totals"]["committed_fy26"]
    print(f"Extracted {data['metadata']['total_line_items']} line items")
    print(f"Total approved: ${approved:,.2f}")
    print(f"Total committed: ${committed:,.2f}")
    print(f"Written to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
